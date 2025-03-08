import json
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from openpyxl import Workbook,load_workbook


def parse_input(directory: str) -> Optional[List[Tuple[List[Dict], List[Dict], List[Dict], List[Dict], Path]]]:
    directory_path = Path(directory)
    all_datasets = []

    if not directory_path.exists() or not directory_path.is_dir():
        print(f"Error: Directory {directory} does not exist or is not a folder.\n")
        return None

    for file_path in directory_path.rglob("*.json"):
        try:
            data = json.loads(file_path.read_text(encoding="utf-8"))

            tasks = data.get("activities", [])
            relations = data.get("relations", [])
            consumptions = data.get("consumptions", [])
            resources = data.get("resources", [])

            if not all(isinstance(lst, list) for lst in [tasks, relations, consumptions, resources]):
                print(f"Warning: Invalid data format in {file_path}\n")
                continue

            all_datasets.append((tasks, relations, consumptions, resources, file_path))

        except json.JSONDecodeError as e:
            print(f"Error: JSON parsing error in {file_path} - {str(e)}\n")
        except Exception as e:
            print(f"Error: Unable to read {file_path} - {str(e)}\n")

    return all_datasets if all_datasets else None


def export_schedule_to_xlsx(
        variables,
        clauses, 
        status,
        output_file,
        time_solve=0,
        ago_type=None,
        file_name=None,
        problem_field=None):

    headers = ['File Name', 'Problem', 'Type', 'Status', 'Time', 'Variables', 'Clauses']
    output_path = Path(output_file)
    

    if output_path.is_file():
        workbook = load_workbook(output_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers) 

    sheet.append([file_name, problem_field, ago_type, status, time_solve, variables, clauses])

    workbook.save(output_path)


class VariableFactory:
    def __init__(self):
        self.var_count = 1
        self.var_map = {}
    def get_var(self, name):
        if name not in self.var_map:
            self.var_map[name] = self.var_count
            self.var_count += 1
        return self.var_map[name]

    def start(self, task_id, time):
        return self.get_var(f"start_{task_id}_{time}")

    def run(self, task_id, time):
        return self.get_var(f"run_{task_id}_{time}")

    def consume(self, task_id, resource_id, time, index):
        return self.get_var(f"consume_{task_id}_{resource_id}_{time}_{index}")



def convert_json_to_old_2014(json_path):
    json_file = Path(json_path)
    if json_file.exists():
        json_string = json_file.read_text(encoding="utf-8")
        json_data = json.loads(json_string)
    
    lines = []
    
    for task in json_data.get("activities", []):
        lines.append(f"task;{task['id']};{task['duration']};{task['name']}")
    
    # Process relations
    for relation in json_data.get("relations", []):
        lines.append(f"aob;{relation['task_id_1']};{relation['task_id_2']};{relation['relation_type']}")
    
    for consumption in json_data.get("consumptions", []):
        lines.append(f"consumption;{consumption['task_id']};{consumption['resource_id']};{consumption['amount']}")


    for resource in json_data.get("resources", []):
        lines.append(f"resource;{resource['id']};{resource['capacity']};{resource['name']}")
    
    return "\n".join(lines)